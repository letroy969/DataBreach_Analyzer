#!/usr/bin/env python3
"""
Create Excel workbook for Data Breach Insights Report.

This script generates a professional Excel workbook with pivot tables,
charts, and advanced formulas demonstrating data analyst skills.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import argparse
import os
from pathlib import Path

def load_breach_data(csv_file: str) -> pd.DataFrame:
    """Load and clean breach data from CSV."""
    print(f"📊 Loading data from {csv_file}...")
    
    df = pd.read_csv(csv_file)
    
    # Clean and enhance data
    df['breach_date'] = pd.to_datetime(df['breach_date'])
    df['year'] = df['breach_date'].dt.year
    df['month'] = df['breach_date'].dt.month
    df['quarter'] = df['breach_date'].dt.quarter
    df['is_large_breach'] = df['records_exposed'] >= 1000000
    
    # Severity classification
    def classify_severity(records):
        if records <= 1000:
            return 'Low'
        elif records <= 10000:
            return 'Medium'
        elif records <= 100000:
            return 'High'
        elif records <= 1000000:
            return 'Critical'
        else:
            return 'Catastrophic'
    
    df['severity_level'] = df['records_exposed'].apply(classify_severity)
    
    print(f"✅ Loaded {len(df)} records")
    return df

def create_industry_lookup() -> pd.DataFrame:
    """Create industry lookup table."""
    lookup_data = {
        'raw': ['Healthcare', 'Financial', 'Technology', 'Retail', 'Government', 
                'Education', 'Energy', 'Manufacturing', 'Transportation', 'Media'],
        'standard': ['Healthcare', 'Financial', 'Technology', 'Retail', 'Government',
                    'Education', 'Energy', 'Manufacturing', 'Transportation', 'Media'],
        'category': ['Critical Infrastructure', 'Critical Infrastructure', 'Information Technology',
                    'Consumer Services', 'Public Sector', 'Public Sector', 'Critical Infrastructure',
                    'Industrial', 'Critical Infrastructure', 'Consumer Services'],
        'risk_level': ['High', 'High', 'Medium', 'Medium', 'High', 'Medium', 'High', 'Medium', 'High', 'Low']
    }
    return pd.DataFrame(lookup_data)

def create_excel_workbook(df: pd.DataFrame, output_file: str):
    """Create comprehensive Excel workbook."""
    print(f"📝 Creating Excel workbook: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 1. RAW tab - Original data
        print("  📥 Creating RAW tab...")
        df_original = df[['id', 'breach_date', 'name', 'industry', 'country', 
                         'records_exposed', 'breach_type', 'source_url']].copy()
        df_original.to_excel(writer, sheet_name='RAW', index=False)
        
        # 2. CLEAN tab - Enhanced data
        print("  🧹 Creating CLEAN tab...")
        clean_columns = ['id', 'breach_date', 'name', 'industry', 'country', 
                        'records_exposed', 'breach_type', 'source_url', 'year', 
                        'month', 'quarter', 'is_large_breach', 'severity_level']
        df[clean_columns].to_excel(writer, sheet_name='CLEAN', index=False)
        
        # 3. Industry lookup table
        print("  🗺️ Creating industry_map tab...")
        industry_lookup = create_industry_lookup()
        industry_lookup.to_excel(writer, sheet_name='industry_map', index=False)
        
        # 4. Pivot tables
        print("  📊 Creating pivot tables...")
        
        # Breaches by year
        yearly_breaches = df.groupby('year').agg({
            'id': 'count',
            'records_exposed': 'sum'
        }).reset_index()
        yearly_breaches.columns = ['Year', 'Breach_Count', 'Total_Records']
        yearly_breaches.to_excel(writer, sheet_name='PIVOT_BreachesByYear', index=False)
        
        # Industry analysis
        industry_analysis = df.groupby('industry').agg({
            'id': 'count',
            'records_exposed': ['sum', 'mean']
        }).reset_index()
        industry_analysis.columns = ['Industry', 'Breach_Count', 'Total_Records', 'Avg_Records']
        industry_analysis = industry_analysis.sort_values('Total_Records', ascending=False)
        industry_analysis.to_excel(writer, sheet_name='PIVOT_IndustryRecords', index=False)
        
        # Geographic analysis
        geo_analysis = df.groupby('country').agg({
            'id': 'count',
            'records_exposed': 'sum'
        }).reset_index()
        geo_analysis.columns = ['Country', 'Breach_Count', 'Total_Records']
        geo_analysis = geo_analysis.sort_values('Total_Records', ascending=False)
        geo_analysis.to_excel(writer, sheet_name='PIVOT_Geography', index=False)
        
        # Breach type analysis
        type_analysis = df.groupby('breach_type').agg({
            'id': 'count',
            'records_exposed': 'sum'
        }).reset_index()
        type_analysis.columns = ['Breach_Type', 'Count', 'Total_Records']
        type_analysis = type_analysis.sort_values('Count', ascending=False)
        type_analysis.to_excel(writer, sheet_name='PIVOT_BreachTypes', index=False)
        
        # 5. Summary statistics
        print("  📈 Creating summary statistics...")
        summary_stats = {
            'Metric': [
                'Total Breaches',
                'Total Records Exposed',
                'Average Breach Size',
                'Largest Breach',
                'Date Range Start',
                'Date Range End',
                'Unique Industries',
                'Unique Countries',
                'Unique Breach Types'
            ],
            'Value': [
                len(df),
                f"{df['records_exposed'].sum():,}",
                f"{df['records_exposed'].mean():,.0f}",
                f"{df['records_exposed'].max():,}",
                df['breach_date'].min().strftime('%Y-%m-%d'),
                df['breach_date'].max().strftime('%Y-%m-%d'),
                df['industry'].nunique(),
                df['country'].nunique(),
                df['breach_type'].nunique()
            ]
        }
        summary_df = pd.DataFrame(summary_stats)
        summary_df.to_excel(writer, sheet_name='Summary_Stats', index=False)
        
        # 6. Top breaches
        print("  🏆 Creating top breaches list...")
        top_breaches = df.nlargest(20, 'records_exposed')[
            ['name', 'industry', 'country', 'breach_date', 'records_exposed', 'breach_type']
        ].copy()
        top_breaches['breach_date'] = top_breaches['breach_date'].dt.strftime('%Y-%m-%d')
        top_breaches.to_excel(writer, sheet_name='Top_Breaches', index=False)
        
        # 7. Executive Summary
        print("  👔 Creating executive summary...")
        exec_summary = {
            'Key Insights': [
                f"Total of {len(df):,} data breaches analyzed",
                f"Over {df['records_exposed'].sum():,} records exposed",
                f"Average breach size: {df['records_exposed'].mean():,.0f} records",
                f"Largest breach: {df['records_exposed'].max():,} records",
                f"Most affected industry: {df.groupby('industry')['records_exposed'].sum().idxmax()}",
                f"Most common breach type: {df['breach_type'].mode().iloc[0]}"
            ]
        }
        exec_df = pd.DataFrame(exec_summary)
        exec_df.to_excel(writer, sheet_name='Executive_Summary', index=False)
    
    print(f"✅ Excel workbook created successfully: {output_file}")

def add_excel_formatting(output_file: str):
    """Add professional formatting to Excel workbook."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, BarChart, Reference
        
        print("🎨 Adding professional formatting...")
        
        wb = load_workbook(output_file)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0b2948", end_color="0b2948", fill_type="solid")
        accent_fill = PatternFill(start_color="1fb6b6", end_color="1fb6b6", fill_type="solid")
        highlight_fill = PatternFill(start_color="ffb86b", end_color="ffb86b", fill_type="solid")
        
        # Format all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        print("✅ Formatting applied successfully")
        
    except ImportError:
        print("⚠️ openpyxl not available for advanced formatting")
    except Exception as e:
        print(f"⚠️ Error applying formatting: {e}")

def main():
    """Main function to create Excel workbook."""
    parser = argparse.ArgumentParser(description="Create Excel workbook for breach analysis")
    parser.add_argument("--csv", default="data/sample_breaches.csv", help="Input CSV file")
    parser.add_argument("--output", default="excel/breach_analysis.xlsx", help="Output Excel file")
    parser.add_argument("--format", action="store_true", help="Apply professional formatting")
    
    args = parser.parse_args()
    
    # Validate input file
    if not os.path.exists(args.csv):
        print(f"❌ CSV file not found: {args.csv}")
        return 1
    
    # Create output directory
    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    
    # Load data
    df = load_breach_data(args.csv)
    
    # Create workbook
    create_excel_workbook(df, args.output)
    
    # Apply formatting if requested
    if args.format:
        add_excel_formatting(args.output)
    
    print(f"\n🎉 Excel workbook created: {args.output}")
    print("💡 Open in Excel to add pivot tables, charts, and slicers")
    print("📖 See excel/README.md for detailed instructions")
    
    return 0

if __name__ == "__main__":
    exit(main())
